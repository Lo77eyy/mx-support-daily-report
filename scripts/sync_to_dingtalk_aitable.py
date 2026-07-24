#!/usr/bin/env python3
"""
MX Support Daily Report → 钉钉 AI 表格同步脚本

功能：
  1. 读取 MX_Support_Daily_Report.xlsx 的 "Daily Report" 数据
  2. 通过 dws CLI 写入钉钉 AI 表格 "Daily Workload" 数据表
  3. 本月内每日替换：每次运行删除当月已有记录后写入新数据
  4. 跨月保留：自动保留之前月份的最后一天数据
  5. 输出完整日志

用法：
  python sync_to_dingtalk_aitable.py
  python sync_to_dingtalk_aitable.py --excel-path "自定义路径.xlsx"
  python sync_to_dingtalk_aitable.py --dry-run

环境变量（可选）：
  DWS_PATH  - dws 可执行文件路径（默认 dws，即 PATH 中查找）

Windows Task Scheduler 配置：
  操作: 启动程序
  程序: C:\\...\\python.exe
  参数: C:\\...\\sync_to_dingtalk_aitable.py
  起始位置: C:\\...\\（脚本所在目录）
"""

import subprocess
import json
import sys
import os
import logging
import argparse
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl，请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Windows 控制台 UTF-8 支持 ──
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── 常量 ──
MEXICO_TZ = timezone(timedelta(hours=-6))

# 字段映射和 ID 从 config.json 加载 (通过 config_loader)
BASE_ID = ""
TABLE_ID = ""
FIELD_MAP = {}
FORMULA_FIELDS = set()

MAX_RECORDS_PER_BATCH = 100
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = SCRIPT_DIR / "MX_Support_Daily_Report.xlsx"
LOG_FILE = SCRIPT_DIR / "sync_log.txt"


# ── 日志配置 ──
def setup_logging(log_file: Path, verbose: bool = False):
    """配置日志：同时输出到文件和控制台"""
    logger = logging.getLogger("sync")
    logger.setLevel(logging.DEBUG)

    # 文件日志（UTF-8）
    fh = logging.FileHandler(log_file, encoding="utf-8", mode="a")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    ))

    # 控制台日志
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.DEBUG if verbose else logging.INFO)
    ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))

    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


# ── dws CLI 调用 ──
def find_dws() -> str:
    """
    查找 dws 可执行文件，按以下优先级：
    1. DWS_PATH 环境变量
    2. 项目自带的 bin/dws (仓库 bin/ 目录)
    3. PATH 中的 dws
    """
    # 1. 环境变量
    env_path = os.environ.get("DWS_PATH", "")
    if env_path:
        return env_path

    # 2. 项目自带的 dws（仓库根目录 bin/ 下）
    project_bin = SCRIPT_DIR.parent / "bin"
    if sys.platform == "win32":
        bundled = project_bin / "dws.exe"
    else:
        bundled = project_bin / "dws"
    if bundled.exists():
        return str(bundled)

    # 3. PATH 中查找
    return "dws"


def run_dws(args: list, logger: logging.Logger, timeout: int = 120) -> dict | None:
    """执行 dws 命令并返回 JSON 结果"""
    dws_path = find_dws()
    cmd = [dws_path] + args + ["--format", "json"]
    logger.debug(f"执行: {' '.join(cmd[:6])}...")

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=timeout,
            encoding="utf-8", errors="replace"
        )
        if result.returncode != 0:
            logger.error(f"dws 命令失败 (exit {result.returncode}): {result.stderr.strip()}")
            return None
        return json.loads(result.stdout)
    except subprocess.TimeoutExpired:
        logger.error(f"dws 命令超时 ({timeout}s)")
        return None
    except FileNotFoundError:
        logger.error(f"未找到 dws 命令，请运行 setup 脚本安装，或手动下载: "
                      f"https://github.com/DingTalk-Real-AI/dingtalk-workspace-cli/releases")
        return None
    except json.JSONDecodeError as e:
        logger.error(f"JSON 解析失败: {e}")
        logger.debug(f"原始输出: {result.stdout[:300]}")
        return None


def query_all_records(logger: logging.Logger) -> list:
    """查询 AI 表格中所有记录（自动翻页）"""
    result = run_dws([
        "aitable", "record", "query",
        "--base-id", BASE_ID,
        "--table-id", TABLE_ID,
        "--all",
    ], logger)

    if not result:
        logger.error("查询记录失败")
        return []

    # dws-core 返回格式: {"records": [...], "hasMore": false, "pages": N}
    records = result.get("records", [])
    if not records:
        # 兼容可能的嵌套格式
        records = result.get("data", {}).get("records", [])
    logger.info(f"查询到 {len(records)} 条现有记录")
    return records


def delete_records_batch(record_ids: list, logger: logging.Logger) -> bool:
    """批量删除记录（每批最多 100 条）"""
    if not record_ids:
        logger.info("无需删除记录")
        return True

    total = len(record_ids)
    logger.info(f"准备删除 {total} 条记录...")

    for i in range(0, total, MAX_RECORDS_PER_BATCH):
        batch = record_ids[i:i + MAX_RECORDS_PER_BATCH]
        batch_num = i // MAX_RECORDS_PER_BATCH + 1
        total_batches = (total + MAX_RECORDS_PER_BATCH - 1) // MAX_RECORDS_PER_BATCH

        ids_str = ",".join(batch)
        result = run_dws([
            "aitable", "record", "delete",
            "--base-id", BASE_ID,
            "--table-id", TABLE_ID,
            "--record-ids", ids_str,
            "--yes",
        ], logger)

        if result and result.get("status") == "success":
            logger.info(f"  删除批次 {batch_num}/{total_batches}: 成功删除 {len(batch)} 条")
        else:
            logger.error(f"  删除批次 {batch_num}/{total_batches}: 失败")
            return False

    return True


def create_records_batch(records: list, logger: logging.Logger) -> bool:
    """批量创建记录（每批最多 100 条）"""
    if not records:
        logger.info("无需创建记录")
        return True

    total = len(records)
    logger.info(f"准备创建 {total} 条记录...")

    for i in range(0, total, MAX_RECORDS_PER_BATCH):
        batch = records[i:i + MAX_RECORDS_PER_BATCH]
        batch_num = i // MAX_RECORDS_PER_BATCH + 1
        total_batches = (total + MAX_RECORDS_PER_BATCH - 1) // MAX_RECORDS_PER_BATCH

        # 使用 --records-file 避免 Windows 命令行长度限制
        records_file = SCRIPT_DIR / "_temp_records.json"
        with open(records_file, "w", encoding="utf-8") as f:
            json.dump(batch, f, ensure_ascii=False)

        result = run_dws([
            "aitable", "record", "create",
            "--base-id", BASE_ID,
            "--table-id", TABLE_ID,
            "--records-file", str(records_file),
        ], logger)

        # 清理临时文件
        try:
            records_file.unlink()
        except OSError:
            pass

        if result and result.get("status") == "success":
            # 兼容两种响应格式
            created = (
                result.get("data", {}).get("records", [])
                or result.get("records", [])
            )
            logger.info(f"  创建批次 {batch_num}/{total_batches}: 成功创建 {len(batch)} 条")
        else:
            logger.error(f"  创建批次 {batch_num}/{total_batches}: 失败")
            return False

    return True


# ── Excel 读取 ──
def read_excel_data(excel_path: Path, logger: logging.Logger) -> list:
    """
    读取 Excel 的 "Daily Report" 数据表。
    跳过前 3 行（标题、生成信息、空行），第 4 行为表头，之后为数据。
    遇到空行（日期列为 None）则跳过，用于分隔不同日期的数据块。
    跳过末尾的 "Daily Totals" 汇总区域。
    """
    logger.info(f"读取 Excel: {excel_path}")

    if not excel_path.exists():
        logger.error(f"Excel 文件不存在: {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if "Daily Report" not in wb.sheetnames:
        logger.error("未找到 'Daily Report' 工作表")
        wb.close()
        return []

    ws = wb["Daily Report"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        logger.error("数据表行数不足")
        return []

    # 第 4 行（index=3）是表头
    headers = [str(h).strip() if h else "" for h in rows[3]]
    logger.debug(f"表头: {headers}")

    # 提取 "Tickets Escalated" 列名（可能包含换行符）
    for i, h in enumerate(headers):
        if "Tickets Escalated" in h:
            headers[i] = "Tickets Escalated"

    # 构建列名 → 列索引映射
    col_index = {}
    for excel_col_name, field_id in FIELD_MAP.items():
        for i, h in enumerate(headers):
            if excel_col_name in h:
                col_index[excel_col_name] = i
                break
        if excel_col_name not in col_index:
            logger.warning(f"未找到列: {excel_col_name}")

    logger.debug(f"列映射: {col_index}")

    # 读取数据行（从第 5 行开始，index=4）
    data_rows = []
    in_totals_section = False

    for row_idx, row in enumerate(rows[4:], start=5):
        # 检查是否进入汇总区域
        first_val = row[0] if row else None
        if first_val and str(first_val).strip() in ("Daily Totals", "GRAND TOTAL"):
            in_totals_section = True
            continue
        if in_totals_section:
            continue

        # 跳过空行
        if not row or all(v is None for v in row):
            continue

        # 提取数据
        record_cells = {}
        has_date = False

        for excel_col_name, field_id in FIELD_MAP.items():
            if excel_col_name not in col_index:
                continue
            idx = col_index[excel_col_name]
            if idx >= len(row):
                continue
            val = row[idx]
            if val is None:
                continue

            # 类型转换
            if excel_col_name == "Data Retrieval Date":
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                else:
                    val = str(val).strip()
                if val:
                    has_date = True
            elif excel_col_name in ("Needs Follow Up", "Tickets Under Name", "Tickets Escalated"):
                val = str(int(val)) if isinstance(val, (int, float)) else str(val).strip()
            else:
                val = str(val).strip()

            if val:
                record_cells[field_id] = val

        if has_date and record_cells.get(FIELD_MAP["Agent Name"]):
            data_rows.append({"cells": record_cells})

    logger.info(f"从 Excel 读取到 {len(data_rows)} 条数据行")
    return data_rows


# ── 核心同步逻辑 ──
def get_record_month(record: dict) -> str | None:
    """从记录中提取月份（YYYY-MM 格式）"""
    cells = record.get("cells", {})
    date_val = cells.get(FIELD_MAP["Data Retrieval Date"], "")
    if not date_val:
        return None
    # 日期格式: YYYY-MM-DD
    match = re.match(r"(\d{4}-\d{2})", str(date_val))
    return match.group(1) if match else None


def get_current_month() -> str:
    """获取当前墨西哥时区的月份"""
    now = datetime.now(MEXICO_TZ)
    return now.strftime("%Y-%m")


def get_current_date() -> str:
    """获取当前墨西哥时区的日期（YYYY-MM-DD）"""
    now = datetime.now(MEXICO_TZ)
    return now.strftime("%Y-%m-%d")


def sync_to_aitable(excel_path: Path, dry_run: bool, logger: logging.Logger,
                    override_date: str = None) -> bool:
    """
    主同步流程：
    1. 读取 Excel 数据，筛选出指定日期的行
    2. 查询 AI 表格现有记录
    3. 删除当月已有记录（跨月保留上月数据）
    4. 写入当天新数据

    override_date: 可选，格式 YYYY-MM-DD，覆盖当天日期（用于测试或补录）
    """
    current_month = get_current_month()
    today = override_date if override_date else get_current_date()
    # 如果指定了日期，月份也从该日期推导
    if override_date:
        current_month = today[:7]  # YYYY-MM
    logger.info(f"{'='*60}")
    logger.info(f"MX Support Daily Report → 钉钉 AI 表格同步")
    logger.info(f"当前墨西哥日期: {today} (月份: {current_month})")
    logger.info(f"目标 Base: {BASE_ID}")
    logger.info(f"目标 Table: {TABLE_ID} (Daily Workload)")
    logger.info(f"{'='*60}")

    if dry_run:
        logger.info("[DRY RUN] 预览模式，不执行写入操作")

    # ── Step 1: 读取 Excel 并筛选当天数据 ──
    logger.info("\n[Step 1/4] 读取 Excel 数据...")
    all_records = read_excel_data(excel_path, logger)
    if not all_records:
        logger.error("没有可同步的数据")
        return False

    # 统计 Excel 中所有日期
    all_dates = set()
    for rec in all_records:
        d = rec["cells"].get(FIELD_MAP["Data Retrieval Date"], "")
        if d:
            all_dates.add(d)
    logger.info(f"Excel 包含的日期: {sorted(all_dates)}")

    # 筛选当天日期的记录
    new_records = [
        rec for rec in all_records
        if rec["cells"].get(FIELD_MAP["Data Retrieval Date"]) == today
    ]
    logger.info(f"当天 ({today}) 数据: {len(new_records)} 条")

    if not new_records:
        logger.warning(f"Excel 中没有日期 {today} 的数据，请检查 Excel 是否已更新")
        logger.info("Excel 中包含的日期: " + ", ".join(sorted(all_dates)))
        return False

    # 统计各日期记录数
    for d in sorted(all_dates):
        count = sum(1 for rec in all_records
                    if rec["cells"].get(FIELD_MAP["Data Retrieval Date"]) == d)
        marker = " ← 今天" if d == today else ""
        logger.info(f"  {d}: {count} 条{marker}")

    if dry_run:
        logger.info(f"\n[DRY RUN] 将要写入的 {len(new_records)} 条记录:")
        for rec in new_records[:5]:
            cells = rec["cells"]
            logger.info(f"  {cells.get(FIELD_MAP['Agent Name'])} | "
                        f"{cells.get(FIELD_MAP['Data Retrieval Date'])} | "
                        f"FollowUp={cells.get(FIELD_MAP['Needs Follow Up'])} | "
                        f"Tickets={cells.get(FIELD_MAP['Tickets Under Name'])} | "
                        f"Escalated={cells.get(FIELD_MAP['Tickets Escalated'])}")
        if len(new_records) > 5:
            logger.info(f"  ... 还有 {len(new_records) - 5} 条")
        return True

    # ── Step 2: 查询现有记录 ──
    logger.info("\n[Step 2/4] 查询 AI 表格现有记录...")
    existing_records = query_all_records(logger)

    # ── Step 3: 删除需要替换的记录 ──
    logger.info("\n[Step 3/4] 清理需要替换的数据...")
    to_delete_ids = []
    other_ids = []

    if override_date:
        # --date 模式（补录）：只删除同一天的记录，不影响其他日期
        logger.info(f"[补录模式] 只删除日期 {today} 的现有记录，保留其他日期")
        for rec in existing_records:
            rec_date = rec.get("cells", {}).get(FIELD_MAP["Data Retrieval Date"], "")
            rec_id = rec.get("recordId")
            if not rec_id:
                continue
            if rec_date == today:
                to_delete_ids.append(rec_id)
            else:
                other_ids.append(rec_id)
    else:
        # 默认模式（每日同步）：删除当月所有记录，保留历史月份
        for rec in existing_records:
            rec_month = get_record_month(rec)
            rec_id = rec.get("recordId")
            if not rec_id:
                continue
            if rec_month == current_month:
                to_delete_ids.append(rec_id)
            else:
                other_ids.append(rec_id)

    logger.info(f"将删除: {len(to_delete_ids)} 条记录")
    logger.info(f"保留: {len(other_ids)} 条记录")

    # 显示保留的记录月份分布
    month_counts = {}
    for rec in existing_records:
        m = get_record_month(rec)
        if m and m != current_month:
            month_counts[m] = month_counts.get(m, 0) + 1
    if month_counts:
        for m in sorted(month_counts.keys()):
            logger.info(f"  保留 {m}: {month_counts[m]} 条")

    if to_delete_ids:
        if not delete_records_batch(to_delete_ids, logger):
            logger.error("删除记录失败，中止同步")
            return False
    else:
        logger.info("无需删除记录")

    # ── Step 4: 写入当天新数据 ──
    logger.info("\n[Step 4/4] 写入当天新数据...")
    if not create_records_batch(new_records, logger):
        logger.error("写入新数据失败")
        return False

    # ── 完成 ──
    logger.info(f"\n{'='*60}")
    logger.info(f"同步完成!")
    logger.info(f"  写入: {len(new_records)} 条记录 (日期: {today})")
    logger.info(f"  删除: {len(to_delete_ids)} 条旧记录")
    logger.info(f"  保留: {len(other_ids)} 条历史记录")
    logger.info(f"  表格链接: https://docs.dingtalk.com/i/nodes/{BASE_ID}")
    logger.info(f"{'='*60}")
    return True


# ── 入口 ──
def main():
    parser = argparse.ArgumentParser(
        description="MX Support Daily Report → 钉钉 AI 表格同步"
    )
    parser.add_argument(
        "--excel-path",
        type=str,
        default=str(DEFAULT_EXCEL_PATH),
        help=f"Excel 文件路径 (默认: {DEFAULT_EXCEL_PATH})"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="预览模式：只读取和显示数据，不写入 AI 表格"
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="显示详细日志"
    )
    parser.add_argument(
        "--date",
        type=str,
        default=None,
        help="指定同步日期 (YYYY-MM-DD)，默认使用当天墨西哥日期。用于测试或补录历史数据"
    )
    args = parser.parse_args()

    # 从 config.json 加载 AI 表格配置
    from config_loader import load_config
    config = load_config()
    dt_config = config.get("dingtalk", {})

    global BASE_ID, TABLE_ID, FIELD_MAP, FORMULA_FIELDS
    BASE_ID = dt_config.get("aitable_base_id", "")
    TABLE_ID = dt_config.get("aitable_table_id", "")
    FIELD_MAP = dt_config.get("field_map", {})
    formula_ids = dt_config.get("formula_field_ids", [])
    FORMULA_FIELDS = set(formula_ids) if formula_ids else set()

    if not BASE_ID or not TABLE_ID:
        print("ERROR: 请在 config.json 中配置 dingtalk.aitable_base_id 和 dingtalk.aitable_table_id",
              file=sys.stderr)
        sys.exit(1)
    if not FIELD_MAP:
        print("ERROR: 请在 config.json 中配置 dingtalk.field_map", file=sys.stderr)
        sys.exit(1)

    # 配置日志
    logger = setup_logging(LOG_FILE, args.verbose)
    logger.info(f"\n{'─'*40}")
    logger.info(f"同步任务启动 @ {datetime.now(MEXICO_TZ).strftime('%Y-%m-%d %H:%M:%S')} (Mexico Time)")

    excel_path = Path(args.excel_path)
    success = sync_to_aitable(excel_path, args.dry_run, logger, override_date=args.date)

    if success:
        logger.info("任务完成: 成功")
        sys.exit(0)
    else:
        logger.error("任务完成: 失败")
        sys.exit(1)


if __name__ == "__main__":
    main()
