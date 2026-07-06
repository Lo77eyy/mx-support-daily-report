"""
从 mx_daily_report_data.json 生成 MX Support 每日报告 Excel。

用法：
  python create_daily_excel.py --input mx_daily_report_data.json
  python create_daily_excel.py --input mx_daily_report_data.json --output Report.xlsx
"""
import json
import sys
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ── 样式 ──
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
STAT_LABEL   = Font(name="Calibri", size=11, color="333333")
STAT_FONT    = Font(name="Calibri", bold=True, size=11)
THIN_BORDER  = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_FILL     = PatternFill("solid", fgColor="F2F7FB")
DATE_FILL    = PatternFill("solid", fgColor="D6E4F0")
RED_FONT     = Font(name="Calibri", bold=True, size=11, color="CC0000")


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def build_excel(data, output_path):
    wb = Workbook()
    summary = data["summary"]
    daily = data["daily_data"]
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ═══════════════ Sheet 1: Daily Report ═══════════════
    ws = wb.active
    ws.title = "Daily Report"
    ws.sheet_properties.tabColor = "1F4E79"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"].value = f"MX Support Daily Report — {summary['date_range']} (Mexico Time)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:E2")
    ws["A2"].value = f"Generated: {ts} | Total Tickets: {summary['total_tickets']} | Needs Follow-up: {summary['total_follow_up']} | Escalated: {summary['total_escalated']}"
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")

    # Headers
    row = 4
    headers = [
        "Agent Name",
        "Data Retrieval Date",
        "Needs Follow Up",
        "Tickets Under Name",
        "Tickets Escalated\n(original agent)",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(headers))
    ws.row_dimensions[row].height = 30

    # Data rows
    row = 5
    current_date = None
    for entry in daily:
        date_str = entry["date"]

        # Date separator row
        if date_str != current_date:
            if current_date is not None:
                row += 1  # blank row between dates
            current_date = date_str

        ws.cell(row=row, column=1, value=entry["agent_name"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=date_str).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")

        # Needs Follow Up — highlight red if > 0
        fu = entry["needs_follow_up"]
        c = ws.cell(row=row, column=3, value=fu)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
        if fu > 0:
            c.font = RED_FONT

        ws.cell(row=row, column=4, value=entry["tickets_under_name"]).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        esc = entry["tickets_escalated"]
        ws.cell(row=row, column=5, value=esc).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        if row % 2 == 0:
            for ci in range(1, 6):
                ws.cell(row=row, column=ci).fill = ALT_FILL

        row += 1

    # ── Daily Totals ──
    row += 1
    ws.cell(row=row, column=1, value="Daily Totals").font = SECTION_FONT
    row += 1
    tot_headers = ["Date", "Total Follow Up", "Total Tickets", "Total Escalated"]
    for ci, h in enumerate(tot_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, len(tot_headers))
    row += 1

    for date_str in summary["dates"]:
        day_entries = [d for d in daily if d["date"] == date_str]
        tot_fu = sum(d["needs_follow_up"] for d in day_entries)
        tot_tk = sum(d["tickets_under_name"] for d in day_entries)
        tot_es = sum(d["tickets_escalated"] for d in day_entries)

        ws.cell(row=row, column=1, value=date_str).border = THIN_BORDER
        ws.cell(row=row, column=2, value=tot_fu).border = THIN_BORDER
        ws.cell(row=row, column=3, value=tot_tk).border = THIN_BORDER
        ws.cell(row=row, column=4, value=tot_es).border = THIN_BORDER
        for ci in range(1, 5):
            ws.cell(row=row, column=ci).alignment = Alignment(horizontal="center")
            if ci >= 2:
                ws.cell(row=row, column=ci).font = STAT_FONT
        if tot_fu > 0:
            ws.cell(row=row, column=2).font = RED_FONT
        row += 1

    # Grand total
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = STAT_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=summary["total_follow_up"]).font = RED_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3, value=summary["total_tickets"]).font = STAT_FONT
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=4, value=summary["total_escalated"]).font = STAT_FONT
    ws.cell(row=row, column=4).border = THIN_BORDER
    for ci in range(1, 5):
        ws.cell(row=row, column=ci).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 24

    # ═══════════════ Sheet 2: Ticket IDs ═══════════════
    ws2 = wb.create_sheet("Ticket IDs")
    ws2.sheet_properties.tabColor = "2E75B6"

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = f"Ticket IDs by Date & Agent"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 28

    row = 3
    id_headers = ["Date", "Agent Name", "Category", "Ticket IDs"]
    for ci, h in enumerate(id_headers, 1):
        ws2.cell(row=row, column=ci, value=h)
    style_header_row(ws2, row, len(id_headers))
    row += 1

    for entry in daily:
        # Follow-up IDs
        if entry["follow_up_ids"]:
            ws2.cell(row=row, column=1, value=entry["date"]).border = THIN_BORDER
            ws2.cell(row=row, column=2, value=entry["agent_name"]).border = THIN_BORDER
            ws2.cell(row=row, column=3, value="Needs Follow Up").border = THIN_BORDER
            ws2.cell(row=row, column=4, value=", ".join(str(i) for i in entry["follow_up_ids"])).border = THIN_BORDER
            row += 1
        # Escalated IDs
        if entry["escalated_ids"]:
            ws2.cell(row=row, column=1, value=entry["date"]).border = THIN_BORDER
            ws2.cell(row=row, column=2, value=entry["agent_name"]).border = THIN_BORDER
            ws2.cell(row=row, column=3, value="Escalated").border = THIN_BORDER
            ws2.cell(row=row, column=4, value=", ".join(str(i) for i in entry["escalated_ids"])).border = THIN_BORDER
            row += 1
        # All IDs
        ws2.cell(row=row, column=1, value=entry["date"]).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=entry["agent_name"]).border = THIN_BORDER
        ws2.cell(row=row, column=3, value="All Tickets").border = THIN_BORDER
        ws2.cell(row=row, column=4, value=", ".join(str(i) for i in entry["ticket_ids"])).border = THIN_BORDER
        if row % 2 == 0:
            for ci in range(1, 5):
                ws2.cell(row=row, column=ci).fill = ALT_FILL
        row += 1

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 80

    # ── Save ──
    wb.save(output_path)
    print(f"Excel 已保存: {output_path}")


if __name__ == "__main__":
    from pathlib import Path
    _script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i",
                        default=str(_script_dir / "mx_daily_report_data.json"),
                        help="输入 JSON 路径 (默认: 脚本同目录下的 mx_daily_report_data.json)")
    parser.add_argument("--output", "-o",
                        default=str(_script_dir / "MX_Support_Daily_Report.xlsx"),
                        help="输出 Excel 路径 (默认: 脚本同目录下的 MX_Support_Daily_Report.xlsx)")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)
    build_excel(data, args.output)
